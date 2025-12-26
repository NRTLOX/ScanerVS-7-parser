import os
import re
import tempfile
import json
import threading
import requests
import openpyxl
from ttkbootstrap.widgets import Checkbutton
import pandas as pd
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from deep_translator import GoogleTranslator
from tkinterdnd2 import TkinterDnD, DND_FILES
import tkinter as tk
from threading import Lock
from tkinter import StringVar, filedialog
from ttkbootstrap import Style
from ttkbootstrap.widgets import Button, Label, Frame, Progressbar, Combobox

GITHUB_TOKEN = ""  # Будет заполняться из конфига
MAX_WORKERS = 2  # Число потоков для GitHub API
SEARCH_URL = "https://api.github.com/search/repositories"
HEADERS = {"Authorization": f"token {GITHUB_TOKEN}"}

cache_lock = Lock()

CACHE_FILE = "updates_cache.json"
CONFIG_FILE = "config.json"
# --- Парсеры HTML-файлов (oval, astra, fstec) ---

def parse_html_oval(html_path, xlsx_path):
    with open(html_path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Уязвимости"
    ws.append(["CVE_ID", "BDU_ID", "CVE_URL", "BDU_URL"])

    for tr in soup.find_all("tr", class_=re.compile("^resultbad")):
        cve_id, bdu_id, cve_url, bdu_url = "", "", "", ""
        for link in tr.find_all("a"):
            href = link.get("href", "")
            text = link.text.strip()
            if "bdu.fstec.ru" in href:
                bdu_id = text
                bdu_url = href
            elif "cve.mitre.org" in href:
                cve_id = text
                cve_url = href
        ws.append([cve_id, bdu_id, cve_url, bdu_url])

    wb.save(xlsx_path)
    return xlsx_path

def parse_html_astra(html_path, xlsx_path):
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Уязвимости"
    # Добавляем новый столбец "Пакет"
    ws.append(['Номер CVE', 'Связанные идентификаторы (BDU)', 'Пакет', 'Описание', 'Рекомендации (ссылки)', 'Уровень критичности'])

    for h3 in soup.find_all('h3', class_='header'):
        header_text = h3.get_text(strip=True)
        if header_text.startswith('2.4.') and 'CVE-' in header_text:
            cve_number = next((p for p in header_text.split() if p.startswith('CVE-')), None)
            table = h3.find_next_sibling('table', class_='table-vulnerabilities')
            if not table:
                continue

            data = {
                'Номер CVE': cve_number, 
                'Связанные идентификаторы': '', 
                'Пакет': '',  # Добавляем поле для пакета
                'Описание': '', 
                'Рекомендации': [], 
                'Уровень критичности': ''
            }
            inside_recommendations = False

            for row in table.find_all('tr', class_='table-vulnerabilities__row'):
                cells = row.find_all('td', class_='table-vulnerabilities__cell')

                if len(cells) >= 1 and 'рекомендации' in cells[0].get_text(strip=True).lower():
                    inside_recommendations = True
                    continue

                if inside_recommendations:
                    link_tag = row.find('a')
                    if link_tag:
                        data['Рекомендации'].append(link_tag['href'])
                        continue
                    else:
                        inside_recommendations = False

                if len(cells) >= 3:
                    key = cells[0].get_text(strip=True).lower()
                    value = cells[-1].get_text(strip=True)
                    if key == 'связанные идентификаторы':
                        bdu = [x.strip() for x in value.split(',') if x.strip().startswith('BDU:')]
                        data['Связанные идентификаторы'] = ', '.join(bdu)
                    elif key == 'уровень критичности':
                        data['Уровень критичности'] = value
                    elif key == 'описание':
                        data['Описание'] = value
                    elif key == 'по/пакет' or key == 'пакет':  # Добавляем обработку пакета
                        data['Пакет'] = value

            ws.append([
                data['Номер CVE'], 
                data['Связанные идентификаторы'], 
                data['Пакет'],  # Добавляем пакет в строку
                data['Описание'], 
                ', '.join(data['Рекомендации']), 
                data['Уровень критичности']
            ])

    wb.save(xlsx_path)
    return xlsx_path

def parse_html_fstec(html_path, xlsx_path):
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Уязвимости"
    # Добавляем новый столбец "Пакет"
    ws.append(['Номер CVE', 'Связанные идентификаторы (CVE)', 'Пакет', 'Описание', 'Рекомендации (ссылки)', 'Уровень критичности'])

    for h3 in soup.find_all('h3', class_='header'):
        header_text = h3.get_text(strip=True)
        if header_text.startswith('2.4.') and 'BDU:' in header_text:
            bdu_number = next((p for p in header_text.split() if p.startswith('BDU:')), None)
            table = h3.find_next_sibling('table', class_='table-vulnerabilities')
            if not table:
                continue

            data = {
                'Номер CVE': bdu_number, 
                'Связанные идентификаторы': '', 
                'Пакет': '',  # Добавляем поле для пакета
                'Описание': '', 
                'Рекомендации': [], 
                'Уровень критичности': ''
            }
            inside_recommendations = False

            for row in table.find_all('tr', class_='table-vulnerabilities__row'):
                cells = row.find_all('td', class_='table-vulnerabilities__cell')

                if len(cells) >= 1 and 'рекомендации' in cells[0].get_text(strip=True).lower():
                    inside_recommendations = True
                    continue

                if inside_recommendations:
                    link_tag = row.find('a')
                    if link_tag:
                        data['Рекомендации'].append(link_tag['href'])
                        continue
                    else:
                        inside_recommendations = False

                if len(cells) >= 3:
                    key = cells[0].get_text(strip=True).lower()
                    value = cells[-1].get_text(strip=True)
                    if key == 'связанные идентификаторы':
                        cve = [x.strip() for x in value.split(',') if x.strip().startswith('CVE-')]
                        data['Связанные идентификаторы'] = ', '.join(cve)
                    elif key == 'уровень критичности':
                        data['Уровень критичности'] = value
                    elif key == 'описание':
                        data['Описание'] = value
                    elif key == 'по/пакет' or key == 'пакет':  # Добавляем обработку пакета
                        data['Пакет'] = value

            ws.append([
                data['Связанные идентификаторы'], 
                data['Номер CVE'], 
                data['Пакет'],  # Добавляем пакет в строку
                data['Описание'], 
                ', '.join(data['Рекомендации']), 
                data['Уровень критичности']
            ])

    wb.save(xlsx_path)
    return xlsx_path

# --- Вспомогательные функции ---

def collect_rows_by_cve(file_path, cve_column_index):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    rows, seen = [], set()
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        cve = str(row[cve_column_index]).strip() if row[cve_column_index] else ''
        if cve and cve not in seen:
            seen.add(cve)
            rows.append((cve, row))
    return header, rows, seen

def merge_cve_rows(output_file, closed_file=None, packages_file=None, *files):
    """
    Объединяет данные из нескольких файлов Excel с уязвимостями в один файл.
    files: список путей к xlsx файлам (от 1 до 3 штук).
    Ожидаемые форматы столбцов:
      - OVAL-файл: CVE в колонке index=2
      - ASTRA/FSTEC: CVE в колонке index=0, пакет в index=2
    """
    def load_packages_from_file(file_path):
        """Загружает пакеты и их версии из файла"""
        packages = set()
        if not file_path or not os.path.exists(file_path):
            return packages
            
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    parts = line.strip().split(';')
                    if len(parts) >= 1:
                        name = parts[0].split(':')[0]  # Удаляем архитектуру если есть
                        packages.add(name.lower())  # Приводим к нижнему регистру для сравнения
        except Exception as e:
            print(f"Ошибка загрузки файла пакетов: {e}")
        return packages

    def collect_cves_from_file(file_path, cve_col_idx):
        """Собирает все CVE из файла в множество"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        cves = set()
        for row in ws.iter_rows(values_only=True, min_row=2):  # Пропускаем заголовок
            cve = str(row[cve_col_idx]).strip() if row[cve_col_idx] else ''
            if cve:
                cves.add(cve)
        return cves

    if not files:
        raise ValueError("Нет входных файлов для объединения")

    # Загружаем список пакетов для проверки
    packages = load_packages_from_file(packages_file) if packages_file else set()

    # Сначала соберем все CVE из каждого файла
    file_cves = {}
    for file_path in files:
        try:
            # Определяем тип файла по содержимому (не по имени)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            first_row = next(ws.iter_rows(values_only=True))
            
            # Определяем индекс колонки с CVE
            if "Номер CVE" in first_row or "CVE_ID" in first_row:
                cve_col_idx = first_row.index("Номер CVE") if "Номер CVE" in first_row else first_row.index("CVE_ID")
            else:
                # Эвристика: если в первом столбце есть CVE - это ASTRA/FSTEC, иначе OVAL
                sample_cve = next((str(cell).strip() for cell in first_row if str(cell).startswith("CVE-")), None)
                cve_col_idx = 0 if sample_cve else 2
            
            cves = collect_cves_from_file(file_path, cve_col_idx)
            file_cves[file_path] = cves
        except Exception as e:
            print(f"Ошибка при обработке файла {file_path}: {e}")
            continue

    # Собираем все строки из всех файлов
    all_rows = []
    headers = []
    for file_path in files:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers.append(next(ws.iter_rows(values_only=True)))  # Заголовок
            
            # Определяем индекс колонки с CVE для этого файла
            first_row = headers[-1]
            if "Номер CVE" in first_row or "CVE_ID" in first_row:
                cve_col_idx = first_row.index("Номер CVE") if "Номер CVE" in first_row else first_row.index("CVE_ID")
            else:
                sample_cve = next((str(cell).strip() for cell in first_row if str(cell).startswith("CVE-")), None)
                cve_col_idx = 0 if sample_cve else 2
            
            # Определяем индекс колонки с пакетом (если есть)
            pkg_col_idx = None
            if "Пакет" in first_row:
                pkg_col_idx = first_row.index("Пакет")
            elif len(first_row) > 2:  # Эвристика: предполагаем, что пакет в 3-й колонке
                pkg_col_idx = 2
            
            for row in ws.iter_rows(values_only=True, min_row=2):  # Пропускаем заголовок
                all_rows.append((row[cve_col_idx], row, pkg_col_idx))
        except Exception as e:
            print(f"Ошибка при чтении строк из {file_path}: {e}")
            continue

    # Удаление дубликатов
    unique_rows = {}
    for cve, row, pkg_col_idx in all_rows:
        cve = str(cve).strip() if cve else ""
        if cve and cve not in unique_rows:
            unique_rows[cve] = (row, pkg_col_idx)

    # Фильтрация закрытых уязвимостей
    closed_rows = []
    if closed_file and os.path.isfile(closed_file):
        closed_cves = set()
        wb_closed = openpyxl.load_workbook(closed_file)
        for row in wb_closed.active.iter_rows(values_only=True, min_row=2):
            if row[1]:  # предполагаем, что CVE во второй колонке
                closed_cves.add(str(row[1]).strip())
        
        # Разделяем на открытые и закрытые
        open_rows = {}
        for cve, (row, pkg_col_idx) in unique_rows.items():
            if cve in closed_cves:
                closed_rows.append((row, pkg_col_idx))
            else:
                open_rows[cve] = (row, pkg_col_idx)
        unique_rows = open_rows

    # Создаем итоговую книгу
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Открытые CVE"
    
    # Создаем лист для закрытых уязвимостей, если они есть
    if closed_rows:
        ws_closed = wb_out.create_sheet("Закрытые CVE")
    
    # Создаем лист для отсутствующих пакетов
    ws_missing = wb_out.create_sheet("Отсутствуют пакеты")

    # Определяем заголовки на основе первого файла
    if headers:
        base_header = list(headers[0])
        # Удаляем возможные дубликаты в заголовках
        seen = set()
        base_header = [x for x in base_header if not (x in seen or seen.add(x))]
    else:
        base_header = ["Номер CVE", "Связанные идентификаторы", "Пакет", "Описание", "Рекомендации", "Уровень критичности"]
    
    extended_header = base_header + ["ФСТЭК", "Астра", "OVAL"]
    ws_out.append(extended_header)
    if closed_rows:
        ws_closed.append(extended_header)
    ws_missing.append(extended_header)
    
    # Функция для проверки пакета
    def is_package_missing(row, pkg_col_idx):
        if not packages or pkg_col_idx is None or len(row) <= pkg_col_idx:
            return False
        pkg_name = str(row[pkg_col_idx]).strip().lower()
        return pkg_name and pkg_name not in packages

    # Заполняем данные
    for cve, (row, pkg_col_idx) in unique_rows.items():
        # Проверяем наличие пакета
        if is_package_missing(row, pkg_col_idx):
            # Переносим на лист отсутствующих пакетов
            marks = []
            for file_path in files:
                marks.append("+" if cve in file_cves.get(file_path, set()) else "")
            
            if len(files) == 1:
                marks = marks * 3
            elif len(files) == 2:
                marks = marks + [""]
            
            extended_row = list(row) + marks[:3]
            ws_missing.append(extended_row)
            continue
            
        # Обрабатываем как обычно
        marks = []
        for file_path in files:
            marks.append("+" if cve in file_cves.get(file_path, set()) else "")
        
        if len(files) == 1:
            marks = marks * 3
        elif len(files) == 2:
            marks = marks + [""]
        
        extended_row = list(row) + marks[:3]
        ws_out.append(extended_row)
    
    # Заполняем закрытые уязвимости
    if closed_rows:
        for row, pkg_col_idx in closed_rows:
            # Проверяем наличие пакета
            if is_package_missing(row, pkg_col_idx):
                marks = []
                for file_path in files:
                    marks.append("+" if str(row[0]).strip() in file_cves.get(file_path, set()) else "")
                
                if len(files) == 1:
                    marks = marks * 3
                elif len(files) == 2:
                    marks = marks + [""]
                
                extended_row = list(row) + marks[:3]
                ws_missing.append(extended_row)
                continue
                
            cve = str(row[0]).strip() if row[0] else ""
            marks = []
            for file_path in files:
                marks.append("+" if cve in file_cves.get(file_path, set()) else "")
            
            if len(files) == 1:
                marks = marks * 3
            elif len(files) == 2:
                marks = marks + [""]
            
            extended_row = list(row) + marks[:3]
            ws_closed.append(extended_row)
    
    wb_out.save(output_file)
    return len(all_rows) - len(unique_rows), len(closed_rows)

def load_update_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[!] Ошибка загрузки кэша: {e}")
    return {}

def save_update_cache(cache):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[!] Ошибка сохранения кэша: {e}")

def extract_update_number_from_url(url):
    with cache_lock:  # 🔐 Блокируем чтение и обновление кэша
        cache = load_update_cache()

    if url in cache:
        return cache[url]

    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        resp = requests.get(url, timeout=7, headers=headers)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "html.parser")
            title = soup.find("title")
            if title:
                match = re.search(
                    r'оператив[^\d]*обновлен[^\d]*№?\s*([A-ZА-Я\d\-.]+)',
                    title.text,
                    re.IGNORECASE
                )
                if match:
                    number = f"Установить оперативное обновление {match.group(1)}"

                    # 🔄 Перечитываем кэш перед обновлением, чтобы не затереть чужие изменения
                    with cache_lock:
                        updated_cache = load_update_cache()
                        updated_cache[url] = number
                        save_update_cache(updated_cache)

                    return number
    except Exception as e:
        print(f"[!] Ошибка запроса {url}: {e}")

    return url  # если не нашли или ошибка — просто вернуть ссылку

#------ Перевод

def extract_text_from_link_and_translate(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, timeout=10, headers=headers)
        if resp.status_code == 200:
            lines = resp.text.splitlines()
            if len(lines) >= 241:
                raw_line = lines[240]  # строка №241 (индексация с 0)
                soup = BeautifulSoup(raw_line, 'html.parser')
                cleaned = soup.get_text(strip=True)
                translated = GoogleTranslator(source='auto', target='ru').translate(cleaned)
                return translated
    except Exception:
        pass
    return "—"

import os
import re
import tempfile
import threading
import requests
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed

from tkinterdnd2 import TkinterDnD, DND_FILES
import tkinter as tk
from tkinter import StringVar, filedialog
from ttkbootstrap import Style
from ttkbootstrap.widgets import Button, Label, Frame, Progressbar, Combobox


# --- Класс GUI ---
class FileEntry:
    def __init__(self, master, file_path, scan_type, on_change):
        self.frame = Frame(master)
        self.frame.pack(fill="x", pady=2)
        Label(self.frame, text=os.path.basename(file_path), width=40).pack(side="left", padx=5)
        self.var = StringVar(value=scan_type)
        self.combo = Combobox(self.frame, textvariable=self.var, values=["oval", "astra", "fstec"], width=10, state="readonly")
        self.combo.pack(side="left", padx=5)
        self.combo.bind("<<ComboboxSelected>>", lambda e: on_change())
        self.path = file_path
    def get_selected_type(self): return self.var.get()
    def get_path(self): return self.path

class VulnParserApp:
    def __init__(self, root):
        self.save_path_var = StringVar()
        self.root = root
        root.title("⛩ PARSERLOX ⛩")
        self.style = Style("darkly")
        self.file_entries, self.temp_files = [], []
        reset_frame = Frame(root)
        reset_frame.pack(anchor="nw", padx=10, pady=10)
        self.reset_btn = Button(reset_frame, text="🔄", bootstyle="danger", 
                              command=self.reset_all, width=3)
        self.reset_btn.pack(side="left", padx=5)
        
        # Добавляем кнопку для выбора файлов
        self.select_files_btn = Button(reset_frame, text="📁 Выбрать файлы", 
                                      bootstyle="info", 
                                      command=self.select_html_files)
        self.select_files_btn.pack(side="left", padx=5)
        
        self.drop_label = Label(root, text="Перетащи сюда файлы (до 3-х) !!! ДЛЯ HYPRLAND env GDK_BACKEND=x11 <comand>", bootstyle="info")
        self.drop_label.pack(pady=0, ipadx=0, ipady=70, fill="both")
        self.drop_label.drop_target_register(DND_FILES)
        self.drop_label.dnd_bind('Drop', self.handle_drop)
        
        self.files_frame = Frame(root)
        self.files_frame.pack(pady=5, fill="x")
        self.closed_path_var, self.desc_path_var = StringVar(), StringVar()
        self._make_file_selector("Путь для сохранения Excel-файла:", self.save_path_var, self.select_save_path)

        self._make_file_selector("Файл с закрытыми уязвимостями: (необязательно)", self.closed_path_var, self.select_closed_file)
        self._make_file_selector("Файл с описаниями уязвимостей (опционально):", self.desc_path_var, self.select_desc_file)


        self.packages_path_var = StringVar()
        self._make_file_selector("Файл с пакетами (имя;версия):", self.packages_path_var, self.select_packages_file)

        self.start_btn = Button(root, text="🚀 Запустить обработку", bootstyle="success", command=self.start_parsing)
        self.start_btn.pack(pady=10)
        self.report_btn = Button(root, text="📤 Выгрузить отчёт", bootstyle="warning", command=self.export_report)
        self.report_btn.pack(pady=5)
        self.report_btn.config(state="disabled")  # Сначала отключена

        self.progress = Progressbar(root, maximum=100)
        self.progress.pack(fill="x", padx=10)
        self.log_text = tk.Text(root, height=10, state='disabled', bg='#1e1e1e', fg="#d4d4d4", wrap='word')
        self.log_text.pack(fill='both', padx=10, pady=(5,10), expand=True)
        self.log_scroll = tk.Scrollbar(root, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=self.log_scroll.set)
        self.log_scroll.pack(side="right", fill="y")
        self.report_mode = tk.BooleanVar(value=False)
        self.mode_var = StringVar(value="Парсер HTML")
        mode_frame = Frame(root)
        mode_frame.pack(pady=5)
        Label(mode_frame, text="Режим работы:", bootstyle="info").pack(side="left", padx=5)
        self.mode_combo = Combobox(mode_frame, textvariable=self.mode_var, 
                                 values=["Парсер HTML", "Объединить .xlsx", "Поиск CVE на GitHub", "Поиск количеств уязвимостей в пакетах"], 
                                 state="readonly")
        self.mode_combo.pack(side="left")
        self.cve_offline_btn = Button(self.root, text="🔎 Поиск CVE по CPE (dpkg) офлайн", bootstyle="info", command=self.start_cve_offline_search)
        self.cve_offline_btn.pack(pady=5)

        # Добавляем поле для GitHub Token
        self.github_token_var = StringVar()
        self._make_github_token_field()
        self.load_config()

    def select_html_files(self):
        """Ручной выбор HTML-файлов через диалоговое окно"""
        if len(self.file_entries) >= 3:
            self.log("⚠️ Максимум 3 HTML-файла")
            return
            
        files = filedialog.askopenfilenames(
            title="Выберите HTML файлы",
            filetypes=[("HTML files", "*.html *.htm"), ("All files", "*.*")]
        )
        
        if not files:
            return
            
        for path in files:
            if len(self.file_entries) >= 3:
                self.log("⚠️ Максимум 3 HTML-файла")
                break
                
            fname = os.path.basename(path).lower()
            if fname.endswith(".html") or fname.endswith(".htm"):
                scan_type = self.detect_type(path)
                entry = FileEntry(self.files_frame, path, scan_type, self.refresh)
                self.file_entries.append(entry)
                self.log(f"📥 Добавлен HTML-файл: {fname}")
            else:
                self.log(f"⚠️ Файл не является HTML: {fname}")
        
        self.refresh()

    def select_packages_file(self):
        path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if path:
            self.packages_path_var.set(path)
            self.log(f"📦 Выбран файл с пакетами: {os.path.basename(path)}")

    def _make_file_selector(self, label_text, var, command):
        frame = Frame(self.root)
        frame.pack(pady=5, fill="x")
        Label(frame, text=label_text, bootstyle="info").pack(side="left", padx=5)
        Label(frame, textvariable=var, width=50, anchor="w", relief="sunken").pack(side="left", padx=5)
        Button(frame, text="Выбрать файл", bootstyle="secondary", command=command).pack(side="left", padx=5)

    def log(self, msg):
        self.log_text.config(state='normal')
        self.log_text.insert('end', msg + "\n")
        self.log_text.see('end')
        self.log_text.config(state='disabled')
        self.root.update_idletasks()

    def handle_drop(self, event):
        paths = self.root.tk.splitlist(event.data)
        for path in paths:
            fname = os.path.basename(path).lower()
            if fname.endswith(".html") or fname.endswith(".htm"):
                if len(self.file_entries) >= 3:
                    self.log("⚠️ Максимум 3 HTML-файла")
                    continue
                scan_type = self.detect_type(path)
                entry = FileEntry(self.files_frame, path, scan_type, self.refresh)
                self.file_entries.append(entry)
                self.log(f"📥 Добавлен HTML-файл: {fname}")
            elif any(x in fname for x in ["закрыт", "closed", "fix"]):
                self.closed_path_var.set(path)
                self.log(f"📥 Файл закрытых уязвимостей: {fname}")
            elif any(x in fname for x in ["описан", "desc", "vullist"]):
                self.desc_path_var.set(path)
                self.log(f"📥 Файл описаний уязвимостей: {fname}")
            else:
                self.log(f"⚠️ Неизвестный файл: {fname}")
        self.refresh()

    def detect_type(self, path):
        name = os.path.basename(path).lower()
        if "fstec" in name: return "fstec"
        elif "astra" in name: return "astra"
        return "oval"

    def refresh(self):
        self.progress['value'] = 0

    def select_closed_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")])
        if path:
            self.closed_path_var.set(path)
            self.log(f"📂 Выбран файл закрытых уязвимостей: {os.path.basename(path)}")

    def select_desc_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")])
        if path:
            self.desc_path_var.set(path)
            self.log(f"📂 Выбран файл описаний: {os.path.basename(path)}")

    def _make_github_token_field(self):
        frame = Frame(self.root)
        frame.pack(pady=5, fill="x")
        Label(frame, text="GitHub Token:", bootstyle="info").pack(side="left", padx=5)
        entry = tk.Entry(frame, textvariable=self.github_token_var, width=50, show="*")
        entry.pack(side="left", padx=5)
        Button(frame, text="Сохранить", bootstyle="secondary", 
              command=self.save_github_token).pack(side="left", padx=5)
    
    def save_github_token(self):
        global GITHUB_TOKEN, HEADERS
        token = self.github_token_var.get().strip()
        if token:
            GITHUB_TOKEN = token
            HEADERS = {"Authorization": f"token {GITHUB_TOKEN}"}
            self.log("✅ GitHub Token сохранен")
            self.save_config()
            
    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.save_path_var.set(data.get("save_path", ""))
                    self.github_token_var.set(data.get("github_token", ""))
                    if self.github_token_var.get():
                        self.save_github_token()  # Обновляем глобальные переменные
            except Exception as e:
                self.log(f"⚠️ Не удалось загрузить config.json: {e}")
    
    def save_config(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({
                    "save_path": self.save_path_var.get(),
                    "github_token": self.github_token_var.get()
                }, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.log(f"⚠️ Не удалось сохранить config.json: {e}")
    
    # Добавляем новую функцию для поиска CVE
    def github_repo_search_count(self, query):
        """Вернуть количество репозиториев по запросу"""
        try:
            response = requests.get(SEARCH_URL, headers=HEADERS, params={"q": query})
            response.raise_for_status()
            return response.json().get("total_count", 0)
        except Exception as e:
            self.log(f"❌ Ошибка при запросе '{query}': {e}")
            return 0
    
    def process_cve(self, cve):
        """Обработать один CVE: найти total_count по запросу"""
        query = f"{cve}"
        total = self.github_repo_search_count(query)
        self.log(f"{cve}: {total} результатов")
        if total > 1:
            return cve, total
        return None
    
    def search_cve_on_github(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Выберите файл с CVE"
        )
        if not file_path:
            return
            
        output_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            title="Сохранить результаты поиска"
        )
        if not output_path:
            return
            
        try:
            df = pd.read_excel(file_path)
            cve_list = df.iloc[:, 0].dropna().astype(str).tolist()
            found = {}
            
            self.log(f"🔍 Начинаю поиск {len(cve_list)} CVE на GitHub...")
            self.progress['value'] = 0
            step = 100 / len(cve_list)
            
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = [executor.submit(self.process_cve, cve) for cve in cve_list]
                
                for future in as_completed(futures):
                    result = future.result()
                    if result:
                        cve, total = result
                        found[cve] = total
                    self.progress['value'] += step
                    self.root.update()
            
            with open(output_path, "w") as f:
                for cve, total in found.items():
                    f.write(f"{cve}: {total} результатов\n")
            
            self.log(f"\n✅ Готово! Найдено {len(found)} CVE. Сохранено в {output_path}")
            self.progress['value'] = 100
            
        except Exception as e:
            self.log(f"❌ Ошибка: {e}")
            self.progress['value'] = 0
    
    def start_parsing(self):
        selected_mode = self.mode_var.get()
        if selected_mode == "Парсер HTML":
            threading.Thread(target=self._run, daemon=True).start()
        elif selected_mode == "Объединить .xlsx":
            threading.Thread(target=self.merge_xlsx_folder, daemon=True).start()
        elif selected_mode == "Поиск CVE на GitHub":
            threading.Thread(target=self.search_cve_on_github, daemon=True).start()
        elif selected_mode == "Поиск количеств уязвимостей в пакетах":
            threading.Thread(target=self.search_vuln_by_packages, daemon=True).start()


    def _run(self):
        if not self.file_entries:
            self.log("❌ Нет HTML-файлов")
            return

        self.progress['value'] = 0
        self.temp_files.clear()
        step = 100 / len(self.file_entries)

        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = []
            for entry in self.file_entries:
                scan_type, path = entry.get_selected_type(), entry.get_path()
                func = {"oval": parse_html_oval, "astra": parse_html_astra, "fstec": parse_html_fstec}.get(scan_type)
                if not func:
                    self.log(f"❌ Нет обработчика для {scan_type}")
                    return
                self.log(f"🔍 Обработка {os.path.basename(path)} как {scan_type}...")
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                tmp.close()
                self.temp_files.append(tmp.name)
                try:
                    func(path, tmp.name)
                    self.log(f"✅ Обработан {os.path.basename(path)}")
                except Exception as e:
                    self.log(f"❌ Ошибка в {path}: {e}")
                    return
                self.progress['value'] += step


        self.log("🔀 Объединяю результаты...")
        out_file = self.save_path_var.get().strip() or os.path.join(os.path.expanduser("~"), "merged_cves.xlsx")
        dup, closed = merge_cve_rows(
        out_file,
        self.closed_path_var.get() or None,
        self.packages_path_var.get() or None,
        *self.temp_files  # передаём столько файлов, сколько есть (1–3)
        )   
        self.log(f"✅ Удалено дубликатов: {dup}, закрытых: {closed}")
        self.progress['value'] = 80

        desc_file = self.desc_path_var.get() or None
        if desc_file and os.path.isfile(desc_file):
            self.log("🔗 Добавляю номера обновлений из wiki-ссылок...")
            try:
                self.add_links_to_merged(out_file, desc_file)
                self.progress['value'] = 100
            except Exception as e:
                self.log(f"❌ Ошибка при добавлении ссылок: {e}")
                self.progress['value'] = 100
        else:
            self.progress['value'] = 100
            self.log(f"📄 Итоговый файл: {out_file}")
        self.report_btn.config(state="normal")
        self.last_out_file = out_file
    def export_report(self):
        path = self.last_out_file.replace(".xlsx", "_FULL.xlsx")

        if not os.path.isfile(path):
            self.log("⚠️ Итоговый файл с обновлениями не найден.")
            return

        try:
            self.log("📥 Загружаем файл с уязвимостями...")
            df = pd.read_excel(path)
            self.log(f"📊 Столбцы в файле: {df.columns.tolist()}")
            self.log(f"📈 Всего строк до фильтрации: {len(df)}")

            # Показываем окно ввода названия компонента и версии системы
            comp_win = tk.Toplevel(self.root)
            comp_win.title("Параметры отчёта")
            comp_win.geometry("400x350")
            
            # Поле для названия компонента
            Label(comp_win, text="Введите название компонента:", bootstyle="info").pack(pady=5)
            comp_var = StringVar()
            comp_entry = tk.Entry(comp_win, textvariable=comp_var, width=40)
            comp_entry.pack(pady=5)
            
            # Поле для версии системы
            Label(comp_win, text="Введите версию системы (например 1.7, опционально):", bootstyle="info").pack(pady=5)
            version_var = StringVar()
            version_entry = tk.Entry(comp_win, textvariable=version_var, width=40)
            version_entry.pack(pady=5)
            
            # Чекбокс для фильтрации по версии
            filter_var = tk.BooleanVar(value=False)
            filter_check = Checkbutton(comp_win, text="Фильтровать по версии системы", variable=filter_var, bootstyle="info")
            filter_check.pack(pady=5)
            
            # Добавляем чекбоксы для фильтрации CVSS
            Label(comp_win, text="Фильтрация по CVSS 3.0:", bootstyle="info").pack(pady=(10,5))
            
            # Создаем переменные для чекбоксов CVSS
            remove_avn_var = tk.BooleanVar(value=False)
            remove_uir_var = tk.BooleanVar(value=False)
            remove_prn_var = tk.BooleanVar(value=False)
            
            # Чекбоксы CVSS
            Checkbutton(comp_win, text="Удалить AV:N", variable=remove_avn_var, 
                    bootstyle="info").pack(anchor="w", padx=20)
            Checkbutton(comp_win, text="Удалить UI:R", variable=remove_uir_var, 
                    bootstyle="info").pack(anchor="w", padx=20)
            Checkbutton(comp_win, text="Удалить PR:N", variable=remove_prn_var, 
                    bootstyle="info").pack(anchor="w", padx=20)
            
            comp_entry.focus()
            done = threading.Event()

            def submit_component():
                done.set()
                comp_win.destroy()

            Button(comp_win, text="OK", command=submit_component, bootstyle="success").pack(pady=10)
            comp_win.grab_set()
            self.root.wait_window(comp_win)

            done.wait()
            component_name = comp_var.get().strip()
            system_version = version_var.get().strip()
            filter_by_version = filter_var.get()
            
            # Логируем выбранные параметры
            self.log(f"⚙️ Параметры фильтрации:")
            self.log(f"   Компонент: {component_name}")
            self.log(f"   Версия системы: {system_version}")
            self.log(f"   Фильтр по версии: {filter_by_version}")
            self.log(f"   Удалить AV:N: {remove_avn_var.get()}")
            self.log(f"   Удалить UI:R: {remove_uir_var.get()}")
            self.log(f"   Удалить PR:N: {remove_prn_var.get()}")
            
            if not component_name:
                self.log("⚠️ Компонент не введён. Отмена.")
                return
                
            if filter_by_version and not system_version:
                self.log("⚠️ Выбрана фильтрация по версии, но версия не введена. Отмена.")
                return

            # Применяем фильтрацию CVSS если есть столбец с вектором
            if "CVSS 3.0" in df.columns:
                self.log("🔍 Начинаем фильтрацию по CVSS 3.0...")
                original_count = len(df)
                
                # Проверяем наличие данных в столбце
                cvss_non_null = df["CVSS 3.0"].notna().sum()
                self.log(f"   Непустых значений CVSS 3.0: {cvss_non_null}/{original_count}")
                
                if cvss_non_null > 0:
                    # Выводим примеры значений для отладки
                    sample_values = df["CVSS 3.0"].dropna().head(3).tolist()
                    self.log(f"   Примеры значений CVSS: {sample_values}")
                
                # Создаем маску для фильтрации
                mask = pd.Series([True] * len(df), index=df.index)
                
                # Проверяем каждый чекбокс и обновляем маску
                if remove_avn_var.get():
                    avn_count_before = len(df[df["CVSS 3.0"].str.contains(r'\bAV:N\b', na=False, regex=True)])
                    mask &= ~df["CVSS 3.0"].str.contains(r'\bAV:N\b', na=False, regex=True)
                    avn_count_after = len(df[~mask])
                    self.log(f"   AV:N: найдено {avn_count_before}, будет удалено {avn_count_after}")
                
                if remove_uir_var.get():
                    uir_count_before = len(df[df["CVSS 3.0"].str.contains(r'\bUI:R\b', na=False, regex=True)])
                    mask &= ~df["CVSS 3.0"].str.contains(r'\bUI:R\b', na=False, regex=True)
                    uir_count_after = len(df[~mask])
                    self.log(f"   UI:R: найдено {uir_count_before}, будет удалено {uir_count_after}")
                
                if remove_prn_var.get():
                    prn_count_before = len(df[df["CVSS 3.0"].str.contains(r'\bPR:N\b', na=False, regex=True)])
                    mask &= ~df["CVSS 3.0"].str.contains(r'\bPR:N\b', na=False, regex=True)
                    prn_count_after = len(df[~mask])
                    self.log(f"   PR:N: найдено {prn_count_before}, будет удалено {prn_count_after}")
                
                # Применяем фильтр
                df_filtered = df[mask]
                removed_count = original_count - len(df_filtered)
                
                self.log(f"   Результат фильтрации: удалено {removed_count} строк")
                self.log(f"   Осталось строк: {len(df_filtered)}")
                
                if removed_count > 0:
                    # Показываем какие уязвимости были удалены
                    removed_rows = df[~mask]
                    if not removed_rows.empty:
                        self.log("   Удаленные уязвимости:")
                        for idx, row in removed_rows.head(5).iterrows():
                            cve = row.get("Номер CVE", row.get("CVE_ID", "N/A"))
                            cvss = row.get("CVSS 3.0", "N/A")
                            self.log(f"     - {cve}: {cvss}")
                        if len(removed_rows) > 5:
                            self.log(f"     ... и еще {len(removed_rows) - 5}")
                
                df = df_filtered
            else:
                self.log("ℹ️ Столбец 'CVSS 3.0' не найден, пропускаем фильтрацию")

            # Фильтрация рекомендаций по версии системы
            if filter_by_version and "Оперативное обновление" in df.columns:
                self.log("🔍 Фильтруем по версии системы...")
                def filter_updates(update_text):
                    if not isinstance(update_text, str):
                        return False
                    
                    if "Обновить ОС до версии" in update_text:
                        return True
                    
                    version_pattern = re.compile(r'оператив[^\d]*обновлен[^\d]*№?\s*([\d.]+)', re.IGNORECASE)
                    match = version_pattern.search(update_text)
                    if match:
                        update_version = match.group(1)
                        return update_version.startswith(system_version + '.') or update_version == system_version
                    return False
                    
                df = df[df["Оперативное обновление"].apply(filter_updates) | 
                    (df["Оперативное обновление"].isna())]
                
                self.log(f"   После фильтрации по версии: {len(df)} строк")

            # Подготовка столбцов
            self.log("📋 Подготавливаем столбцы для отчета...")
            if "Описание" in df.columns and "Связанные идентификаторы (BDU)" in df.columns:
                df["Связанные идентификаторы (BDU)"] = df["Описание"]
                self.log("   Скопировали Описание в Связанные идентификаторы")

            if "Уровень критичности" in df.columns:
                df.rename(columns={"Уровень критичности": "Критичность"}, inplace=True)
                self.log("   Переименовали Уровень критичности в Критичность")

            if "Оперативное обновление" in df.columns:
                df.rename(columns={"Оперативное обновление": "Рекомендации по устранению"}, inplace=True)
                self.log("   Переименовали Оперативное обновление")

            rename_map = {
                "Номер CVE": "Идентификатор уязвимости",
                "Связанные идентификаторы (BDU)": "Описание",
                "Критичность": "Критичность",
                "Рекомендации (ссылки)": "Компонент"
            }
            
            selected_cols = [col for col in rename_map if col in df.columns]
            self.log(f"   Выбранные колонки: {selected_cols}")
            
            df = df[selected_cols + (["Рекомендации по устранению"] if "Рекомендации по устранению" in df.columns else [])]
            df.rename(columns=rename_map, inplace=True)

            if "Компонент" in df.columns and len(df) > 1:
                df.loc[0:, "Компонент"] = component_name
                self.log(f"   Заполнили компонентом '{component_name}'")

            if len(df.columns) >= 3:
                report_path = path.replace("_FULL.xlsx", " - К ОТЧЕТУ.xlsx")
                df.to_excel(report_path, index=False)
                self.log(f"✅ Отчётный файл создан: {report_path}")
                self.log(f"📊 Итоговое количество строк: {len(df)}")
            else:
                self.log(f"⚠️ Недостаточно колонок для отчёта. Получено: {df.columns.tolist()}")

        except Exception as e:
            self.log(f"❌ Ошибка при создании отчёта: {e}")
            import traceback
            self.log(f"❌ Трассировка: {traceback.format_exc()}")




    def add_links_to_merged(self, merged_path, desc_path):
        df_merged = pd.read_excel(merged_path)
        column3 = df_merged.iloc[:, 2]
        self.log("🌐 Перевожу описания из ссылок в 3 столбце...")

        def process_desc_link(index_val):
            i, val = index_val
            if isinstance(val, str) and val.startswith("http"):
                self.log(f"🔎 {i+1}: Парсинг {val}")
                return i, extract_text_from_link_and_translate(val)
            else:
                return i, val

        translated_descriptions = [None] * len(column3)
        with ThreadPoolExecutor(max_workers=20) as executor:
            for i, result in executor.map(process_desc_link, enumerate(column3)):
                translated_descriptions[i] = result

        df_merged.iloc[:, 2] = translated_descriptions

        
        df_desc = pd.read_excel(desc_path)

        file1_vals = df_merged.iloc[:, 1]  # предполагаем, что тут BDU или CVE ID
        file2_keys = df_desc.iloc[:, 0]    # предполагаем, что тут ключи
        file2_targets = df_desc.iloc[:, 13]  # предполагаем, что тут ссылки

        value_to_link = dict(zip(file2_keys, file2_targets))

        urls = []
        for val in file1_vals:
            raw_link = value_to_link.get(val, "")
            match = re.search(r'(https?://wiki\.astralinux[^\s"]+)', str(raw_link))
            url = match.group(1) if match else ""
            urls.append(url)

        found_urls = [url for url in urls if url]
        self.log(f"🔗 Найдено ссылок на wiki.astra: {len(found_urls)}")

        results = ["—"] * len(urls)

        with ThreadPoolExecutor(max_workers=30) as executor:
            futures = {executor.submit(extract_update_number_from_url, url): i for i, url in enumerate(urls) if url}
            for future in as_completed(futures):
                i = futures[future]
                url = urls[i]
                self.log(f"🌐 Парсим: {url}")
                try:
                    results[i] = future.result()
                except Exception as e:
                    self.log(f"⚠️ Ошибка парсинга {url}: {e}")
                    results[i] = "—"

        df_merged['Оперативное обновление'] = [r if isinstance(r, str) else str(r) for r in results]
        out_path = merged_path.replace(".xlsx", "_FULL.xlsx")
        df_merged.to_excel(out_path, index=False)
        self.log(f"📄 Файл сохранён: {out_path}")
    
    def select_save_path(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
            title="Выберите путь для сохранения результата"
        )
        if path:
            self.save_path_var.set(path)
            self.log(f"💾 Файл будет сохранён как: {os.path.basename(path)}")
        self.save_config()


    

    def merge_xlsx_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с .xlsx файлами")
        if not folder:
            self.log("⚠️ Папка не выбрана.")
            return

        self.log(f"📁 Объединяю .xlsx из папки: {folder}")
        try:
            all_data = []  # Будем хранить все данные
            file_paths = []  # Для отслеживания источников
            
            # Читаем все файлы
            for filename in os.listdir(folder):
                if filename.endswith(".xlsx"):
                    path = os.path.join(folder, filename)
                    self.log(f"📄 Читаю файл: {filename}")
                    
                    try:
                        df = pd.read_excel(path)
                        # Добавляем столбец с именем файла для отладки
                        df['source_file'] = filename
                        all_data.append(df)
                        file_paths.append(path)
                    except Exception as e:
                        self.log(f"⚠️ Ошибка чтения файла {filename}: {e}")
                        continue
            
            if not all_data:
                self.log("❌ Нет .xlsx файлов в папке.")
                return
            
            self.log(f"✅ Прочитано файлов: {len(all_data)}")
            
            # Объединяем все данные
            merged_df = pd.concat(all_data, ignore_index=True)
            
            # Проверяем структуру данных
            self.log(f"📊 Столбцы в объединенном датафрейме: {list(merged_df.columns)}")
            self.log(f"📈 Всего строк до обработки: {len(merged_df)}")
            
            if len(merged_df.columns) < 4:
                self.log("❌ Файлы должны содержать минимум 4 столбца")
                return
            
            # Проверяем названия столбцов
            first_col_name = merged_df.columns[0]
            fourth_col_name = merged_df.columns[3] if len(merged_df.columns) > 3 else "Компонент"
            
            self.log(f"🔍 Первый столбец: '{first_col_name}', Четвертый столбец: '{fourth_col_name}'")
            
            # Проверяем наличие дубликатов по первому столбцу
            duplicates = merged_df.duplicated(subset=[first_col_name], keep=False)
            duplicate_count = duplicates.sum()
            
            if duplicate_count > 0:
                self.log(f"🔍 Найдено дубликатов по первому столбцу: {duplicate_count}")
                
                # Создаем словарь для хранения данных
                result_rows = {}
                
                # Обрабатываем каждую строку
                for idx, row in merged_df.iterrows():
                    key = str(row[first_col_name]).strip() if pd.notna(row[first_col_name]) else f"empty_{idx}"
                    
                    if key not in result_rows:
                        # Первое вхождение - сохраняем всю строку
                        result_rows[key] = {
                            'row': row.tolist(),
                            'components': {str(row[fourth_col_name]).strip()} if pd.notna(row[fourth_col_name]) else set(),
                            'files': {row['source_file']}
                        }
                    else:
                        # Дубликат - объединяем компоненты
                        if pd.notna(row[fourth_col_name]):
                            component = str(row[fourth_col_name]).strip()
                            result_rows[key]['components'].add(component)
                        result_rows[key]['files'].add(row['source_file'])
                
                # Формируем новый датафрейм
                processed_rows = []
                for key, data in result_rows.items():
                    row_data = data['row'][:4]  # Первые 4 столбца
                    
                    # Объединяем компоненты через запятую
                    if data['components']:
                        combined_components = ', '.join(sorted(data['components']))
                    else:
                        combined_components = ''
                    
                    # Заменяем четвертый столбец на объединенные компоненты
                    if len(row_data) > 3:
                        row_data[3] = combined_components
                    
                    # Добавляем остальные столбцы (если есть)
                    if len(data['row']) > 4:
                        row_data.extend(data['row'][4:])
                    
                    processed_rows.append(row_data)
                
                # Создаем новый датафрейм
                columns = list(merged_df.columns)
                result_df = pd.DataFrame(processed_rows, columns=columns)
                
            else:
                self.log("ℹ️ Дубликатов не найдено, оставляю данные как есть")
                result_df = merged_df.drop(columns=['source_file'])  # Удаляем служебный столбец
            
            # Сохраняем результат
            out_path = os.path.join(folder, "merged_deduplicated.xlsx")
            
            # Удаляем служебный столбец если он есть
            if 'source_file' in result_df.columns:
                result_df = result_df.drop(columns=['source_file'])
            
            result_df.to_excel(out_path, index=False)
            
            self.log(f"✅ Результат сохранён: {out_path}")
            self.log(f"📈 Итоговое количество строк: {len(result_df)}")
            
            # Показываем пример результата
            if len(result_df) > 0:
                self.log("📋 Пример результата (первые 3 строки):")
                for i in range(min(3, len(result_df))):
                    row = result_df.iloc[i]
                    cve = row.iloc[0] if pd.notna(row.iloc[0]) else "N/A"
                    component = row.iloc[3] if len(row) > 3 and pd.notna(row.iloc[3]) else "N/A"
                    self.log(f"  {i+1}. {cve} -> Компоненты: {component}")
            
        except Exception as e:
            self.log(f"❌ Ошибка объединения: {e}")
            import traceback
            self.log(f"❌ Трассировка: {traceback.format_exc()}")

    def parse_dpkg_line(self, line):
        parts = line.strip().split()
        if len(parts) >= 3 and parts[0] == 'ii':
            pkg = parts[1]
            ver_full = parts[2]
            ver = ver_full.split('-')[0]
            # Формируем упрощенный CPE-строку (вендор не указывается, ставим '*')
            return f"cpe:2.3:a:*:{pkg}:{ver}"
        return None

    def search_cves_in_file(self, cpe_list, json_path):
        matches = []
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for item in data.get("CVE_Items", []):
                nodes = item.get("configurations", {}).get("nodes", [])
                for node in nodes:
                    for cpe_match in node.get("cpe_match", []):
                        cpe_uri = cpe_match.get("cpe23Uri", "")
                        for cpe_query in cpe_list:
                            if cpe_query in cpe_uri:
                                cve_id = item["cve"]["CVE_data_meta"]["ID"]
                                matches.append((cpe_query, cve_id))
                                break
        except Exception as e:
            self.log(f"❌ Ошибка чтения {json_path}: {e}")
        return matches

    def start_cve_offline_search(self):
        # Запускается в отдельном потоке из GUI
        threading.Thread(target=self.cve_offline_search, daemon=True).start()

    def cve_offline_search(self):
        self.log("📂 Выберите файл с выводом dpkg -l")
        dpkg_path = filedialog.askopenfilename(
            title="Выберите файл dpkg -l",
            filetypes=[("Текстовые файлы", "*.txt *.log *.out"), ("Все файлы", "*.*")]
        )
        if not dpkg_path:
            self.log("❗ Файл dpkg не выбран, отмена.")
            return

        self.log("🔍 Читаю dpkg -l и формирую CPE...")
        with open(dpkg_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        cpe_list = list(filter(None, (self.parse_dpkg_line(line) for line in lines)))
        self.log(f"✅ Найдено {len(cpe_list)} CPE")

        self.log("📂 Выберите JSON файлы с базой CVE")
        json_paths = filedialog.askopenfilenames(
            title="Выберите JSON файлы CVE",
            filetypes=[("JSON файлы", "*.json"), ("Все файлы", "*.*")]
        )
        if not json_paths:
            self.log("❗ JSON файлы не выбраны, отмена.")
            return

        self.log(f"⚡ Запускаю поиск CVE в {len(json_paths)} файлах с многопоточностью...")
        matches = []
        with ThreadPoolExecutor(max_workers=50) as executor:
            futures = [executor.submit(self.search_cves_in_file, cpe_list, path) for path in json_paths]
            for future in as_completed(futures):
                res = future.result()
                matches.extend(res)
                self.log(f"🔎 Обработан файл, найдено уязвимостей: {len(res)}")

        if not matches:
            self.log("❌ Уязвимости не найдены.")
            return

        self.log(f"✅ Всего найдено уязвимостей: {len(matches)}")

        save_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Текстовые файлы", "*.txt"), ("Все файлы", "*.*")],
            title="Сохранить результаты поиска"
        )
        if save_path:
            with open(save_path, 'w', encoding='utf-8') as f:
                for cpe, cve in matches:
                    f.write(f"{cpe} → {cve}\n")
            self.log(f"💾 Результаты сохранены в {save_path}")
        else:
            self.log("⚠️ Сохранение отменено.")
    def search_vuln_by_packages(self):
        txt_path = filedialog.askopenfilename(
            title="Выберите .txt файл с пакетами (формат: имя;версия)",
            filetypes=[("Текстовые файлы", "*.txt"), ("Все файлы", "*.*")]
        )
        if not txt_path:
            self.log("❗ Файл с пакетами не выбран.")
            return

        packages = {}
        with open(txt_path, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(';')
                if len(parts) == 2:
                    name, version = parts
                    packages[name.strip()] = version.strip()

        self.log(f"📦 Загружено {len(packages)} пакетов")

        vuln_counts = {pkg: 0 for pkg in packages}

        for entry in self.file_entries:
            scan_type = entry.get_selected_type()
            if scan_type not in ["fstec", "astra"]:
                continue
            try:
                with open(entry.get_path(), encoding='utf-8') as f:
                    soup = BeautifulSoup(f, 'html.parser')

                rows = soup.find_all('tr', class_='table-vulnerabilities__row')

                for row in rows:
                    cols = row.find_all('td')
                    if len(cols) >= 3 and 'ПО/Пакет' in cols[0].get_text(strip=True):
                        pkg_name = cols[2].get_text(strip=True)
                        if pkg_name in packages:
                            vuln_counts[pkg_name] += 1
            except Exception as e:
                self.log(f"❌ Ошибка чтения {entry.get_path()}: {e}")

        save_dir = self.save_path_var.get().strip()
        if not save_dir:
            self.log("❗ Путь для сохранения не указан.")
            return

        out_path = os.path.join(os.path.dirname(save_dir), "уязвимости по пакетам.xlsx")
        df_out = pd.DataFrame(list(vuln_counts.items()), columns=["Пакет", "Количество уязвимостей"])
        df_out.to_excel(out_path, index=False)
        self.log(f"✅ Сохранено: {out_path}")

    def reset_all(self):
        """Сбросить все загруженные файлы и очистить логи"""
        # Удаляем все файловые записи
        for entry in self.file_entries:
            entry.frame.destroy()
        self.file_entries = []
        
        # Очищаем временные файлы
        for tmp_file in self.temp_files:
            try:
                if os.path.exists(tmp_file):
                    os.unlink(tmp_file)
            except Exception:
                pass
        self.temp_files = []
        
        # Сбрасываем пути к файлам
        self.closed_path_var.set("")
        self.desc_path_var.set("")
        
        # Очищаем логи
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', tk.END)
        self.log_text.config(state='disabled')
        
        # Сбрасываем прогресс
        self.progress['value'] = 0
        
        self.log("✅ Все файлы и логи сброшены. Можно загружать новые.")






if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = VulnParserApp(root)
    root.geometry("1400x1000")
    root.mainloop()